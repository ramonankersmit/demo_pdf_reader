from pathlib import Path
import tempfile
import unittest

from pdf_reader.engines.base import Cell, TableExtractionEngine
from openpyxl import load_workbook

from pdf_reader.fallback_import import (
    import_directory_with_fallback,
    import_table_with_fallback,
    table_to_markdown,
)
from pdf_reader.table_extractor import TableExtractor


def build_table(rows: int, cols: int = 2):
    table = []
    for r in range(rows):
        table.append([Cell(text=f"r{r}c{c}") for c in range(cols)])
    return table


class StubEngine(TableExtractionEngine):
    def __init__(self, name, tables=None, should_error=False):
        super().__init__(name)
        self._tables = tables or []
        self._should_error = should_error

    def extract(self, pdf_path: Path):  # type: ignore[override]
        if self._should_error:
            raise RuntimeError("engine failure")
        return self._tables


class StubExtractor(TableExtractor):
    def __init__(self, engines):
        self._engines = {}
        for engine in engines:
            self.register(engine)


class ImportTableWithFallbackTests(unittest.TestCase):
    def test_prefers_first_engine_when_table_has_enough_rows(self):
        first_engine = StubEngine("pymupdf4llm", tables=[build_table(12)])
        backup_engine = StubEngine("pdfplumber", tables=[build_table(15)])
        extractor = StubExtractor([first_engine, backup_engine])

        result = import_table_with_fallback(Path("dummy.pdf"), extractor=extractor)

        self.assertEqual(result.engine, "pymupdf4llm")
        self.assertTrue(result.markdown.startswith("| r0c0 |"))

    def test_falls_back_when_first_engine_insufficient_rows(self):
        first_engine = StubEngine("pymupdf4llm", tables=[build_table(5)])
        backup_engine = StubEngine("pdfplumber", tables=[build_table(11)])
        extractor = StubExtractor([first_engine, backup_engine])

        result = import_table_with_fallback(Path("dummy.pdf"), extractor=extractor)

        self.assertEqual(result.engine, "pdfplumber")
        self.assertIn("r10c1", result.markdown)

    def test_raises_import_error_when_no_engine_succeeds(self):
        first_engine = StubEngine("pymupdf4llm", tables=[])
        backup_engine = StubEngine("pdfplumber", tables=[build_table(3)])
        extractor = StubExtractor([first_engine, backup_engine])

        with self.assertRaises(ImportError):
            import_table_with_fallback(Path("dummy.pdf"), extractor=extractor)


class ImportDirectoryWithFallbackTests(unittest.TestCase):
    def test_writes_excel_and_returns_results(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            base_dir = Path(tmp_dir)
            for filename in ("first.pdf", "second.pdf"):
                (base_dir / filename).write_bytes(b"")

            first_table = build_table(12)
            second_table = build_table(11)

            class PathAwareEngine(TableExtractionEngine):
                def __init__(self, name, mapping):
                    super().__init__(name)
                    self._mapping = mapping

                def extract(self, pdf_path: Path):  # type: ignore[override]
                    return self._mapping.get(Path(pdf_path).name, [])

            first_engine = PathAwareEngine(
                "pymupdf4llm",
                mapping={"first.pdf": [first_table], "second.pdf": [build_table(5)]},
            )
            backup_engine = PathAwareEngine(
                "pdfplumber",
                mapping={"first.pdf": [second_table], "second.pdf": [second_table]},
            )
            extractor = StubExtractor([first_engine, backup_engine])

            output_excel = base_dir / "results.xlsx"

            results = import_directory_with_fallback(base_dir, output_excel, extractor=extractor)

            self.assertIn("first.pdf", results)
            self.assertIn("second.pdf", results)
            self.assertEqual(results["first.pdf"].engine, "pymupdf4llm")
            self.assertEqual(results["second.pdf"].engine, "pdfplumber")

            workbook = load_workbook(output_excel)
            self.assertEqual(len(workbook.sheetnames), 2)
            first_sheet = workbook[workbook.sheetnames[0]]
            self.assertEqual(first_sheet.cell(row=1, column=1).value, "Engine")
            self.assertEqual(first_sheet.cell(row=1, column=2).value, "pymupdf4llm")
            self.assertEqual(first_sheet.cell(row=3, column=1).value, "r0c0")
            self.assertEqual(first_sheet.cell(row=3, column=2).value, "r0c1")

    def test_raises_when_no_pdfs_found(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            empty_dir = Path(tmp_dir)
            with self.assertRaises(ValueError):
                import_directory_with_fallback(empty_dir, empty_dir / "out.xlsx")


class TableToMarkdownTests(unittest.TestCase):
    def test_renders_markdown_with_padding(self):
        table = [
            [Cell("a"), Cell("b"), Cell("c")],
            [Cell("1"), Cell("2")],
        ]
        markdown = table_to_markdown(table)
        expected = "| a | b | c |\n| --- | --- | --- |\n| 1 | 2 |  |"
        self.assertEqual(markdown, expected)


if __name__ == "__main__":
    unittest.main()
